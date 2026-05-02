#!/usr/bin/env python3
"""Read alert.theleetclub.com spec xlsx (values + cell comments). Requires openpyxl."""
import sys

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else "/mnt/c/Users/mahdi/Downloads/alert.theleetclub.com.xlsx"
    wb = openpyxl.load_workbook(path, data_only=False)
    for sn in wb.sheetnames:
        ws = wb[sn]
        mc = ws.max_column or 1
        print("\n### SHEET:", repr(sn), "rows:", ws.max_row, "cols:", mc)
        for r in range(1, (ws.max_row or 0) + 1):
            parts = []
            for c in range(1, mc + 1):
                cell = ws.cell(r, c)
                v = cell.value
                com = ""
                if cell.comment:
                    com = (cell.comment.text or "").strip().replace("\n", " ")
                if v is None and not com:
                    continue
                hdr = openpyxl.utils.get_column_letter(c) + str(r)
                if com:
                    parts.append(f"{hdr}={v!r} NOTE:{com[:200]}")
                else:
                    parts.append(f"{hdr}={v!r}")
            if parts:
                print("ROW", r, ":", " | ".join(parts))


if __name__ == "__main__":
    main()
